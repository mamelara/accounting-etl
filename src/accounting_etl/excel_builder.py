"""Excel spreadsheet builder for output."""

from pathlib import Path
from typing import Dict, List
import pandas as pd
from datetime import datetime
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from accounting_etl.pdf_parser import Transaction


class ExcelBuilder:
    """Builds Excel spreadsheets from transaction data."""

    def build(self, transactions: List[Transaction],
              output_dir: Path,
              funder_codes: Dict[str, str] = None,
              gl_codes: Dict[str, str] = None,
              location_codes: Dict[str, str] = None) -> Path:
        """
        Create Excel file from transactions.

        Args:
            transactions: List of Transaction objects
            output_dir: Directory to save the Excel file
            funder_codes: Dictionary of funder codes {code: description}
            gl_codes: Dictionary of GL codes {code: description}
            location_codes: Dictionary of location codes {code: description}

        Returns path to created file.
        """
        # Convert transactions to DataFrame
        data = []
        for txn in transactions:
            data.append({
                'Date': txn.date,
                'Vendor': txn.vendor,
                'Description': txn.description,
                'G/L Account': txn.gl_account,
                'Location': txn.location,
                'Program': txn.program,
                'Funder': txn.funder,
                'Dept': txn.department,
                'Amount': txn.amount,
                'Receipt_Received': 'No'  # Default to No
            })

        df = pd.DataFrame(data)

        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"credit_card_transactions_{timestamp}.xlsx"
        output_path = output_dir / filename

        # Write to Excel with dropdown validation
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Transactions', index=False)

            worksheet = writer.sheets['Transactions']
            workbook = writer.book

            # Create a hidden sheet for dropdown lists
            if (gl_codes and len(gl_codes) > 0) or \
               (location_codes and len(location_codes) > 0) or \
               (funder_codes and len(funder_codes) > 0):
                lists_sheet = workbook.create_sheet('Dropdown_Lists')
                lists_sheet.sheet_state = 'hidden'

            # Add dropdown validations for each code type
            col_offset = 1
            if gl_codes and len(gl_codes) > 0:
                print(f"  Adding GL Account dropdown to column D ({len(gl_codes)} codes)")
                self._add_dropdown(worksheet, workbook, lists_sheet, gl_codes, len(df), 'D', 'G/L Account', col_offset)
                col_offset += 1
            if location_codes and len(location_codes) > 0:
                print(f"  Adding Location dropdown to column E ({len(location_codes)} codes)")
                self._add_dropdown(worksheet, workbook, lists_sheet, location_codes, len(df), 'E', 'Location', col_offset)
                col_offset += 1
            if funder_codes and len(funder_codes) > 0:
                print(f"  Adding Funder dropdown to column G ({len(funder_codes)} codes)")
                self._add_dropdown(worksheet, workbook, lists_sheet, funder_codes, len(df), 'G', 'Funder', col_offset)

        return output_path

    def _add_dropdown(self, worksheet, workbook, lists_sheet, codes: Dict[str, str],
                      num_rows: int, column: str, column_name: str, col_offset: int):
        """Add dropdown validation to a specific column using a hidden sheet for the list."""
        # Create list of options in format "code - description"
        options = [f"{code} - {desc}" for code, desc in sorted(codes.items())]

        # Write options to hidden sheet in a column
        list_col = get_column_letter(col_offset)
        for idx, option in enumerate(options, start=1):
            lists_sheet[f'{list_col}{idx}'] = option

        # Create reference to the range in the hidden sheet
        list_range = f"Dropdown_Lists!${list_col}$1:${list_col}${len(options)}"

        # Create dropdown validation using the range reference
        dv = DataValidation(
            type="list",
            formula1=list_range,
            allow_blank=True
        )
        dv.error = f'Invalid {column_name}'
        dv.errorTitle = 'Invalid Entry'
        dv.prompt = f'Please select a {column_name} from the dropdown'
        dv.promptTitle = f'{column_name} Selection'

        # Add validation to worksheet
        worksheet.add_data_validation(dv)

        # Apply to all rows in the specified column
        # Start from row 2 (after header) to row num_rows + 1
        for row in range(2, num_rows + 2):
            dv.add(f'{column}{row}')
